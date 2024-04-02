from pickle import TRUE
from django.shortcuts import render, redirect
from django.contrib.auth.views import LoginView
from .forms import *
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import user_passes_test
from .models import *
from django.contrib.auth import logout
import os
from django.http import HttpResponse
from django.db.models import Sum
import datetime
from dateutil.relativedelta import relativedelta
from django.db import connection
import pandas as pd
import numpy as np
from .admin import DemandDataResource  # Import your resource class
from tablib import Dataset
import openpyxl



def import_data_from_excel(file_path):
    print(openpyxl.load_workbook(file_path).sheetnames)
    try:
        # Read the Excel file using pandas

        # Loop through the rows and update existing model instances or create new ones
        if 'DemandData' in openpyxl.load_workbook(file_path).sheetnames:
            df_demanddata = pd.read_excel(file_path, sheet_name='DemandData')
            for index, row in df_demanddata.iterrows():
                gen_station_id = row['GenStationID']
                date = row['Date']
            
                # Check if the record exists, and create it if not
                instance, created = DemandData.objects.get_or_create(GenStationID=gen_station_id, Date=date)
            
                # Update the fields
                instance.GenStationName = row['GenStationName']
                instance.GenType = row['GenType']
                instance.InstalledCap = row['InstalledCap']
                instance.MorningPeak = row['MorningPeak']
                instance.EveningPeak = row['EveningPeak']
                instance.Energy = row['Energy']
                instance.save()
            print('Uploaded Demand data')

        if 'MaxDemandCityandSolar' in openpyxl.load_workbook(file_path).sheetnames:
            df_maxcitysolar = pd.read_excel(file_path, sheet_name='MaxDemandCityandSolar')
            for index, row in df_maxcitysolar.iterrows():
                pid = row['PID']
                date = row['Date']
            
                # Check if the record exists, and create it if not
                instance, created = MaxCitySolar.objects.get_or_create(PID=pid, Date=date)
            
                # Update the fields
                instance.Name = row['Name']
                instance.MaxDemand = row['MaxDemand']
                instance.Time = row['Time']
                instance.save()
            print('Uploaded City and Solar Max data')

        if 'GridFreq' in openpyxl.load_workbook(file_path).sheetnames:
            df_gridfreq = pd.read_excel(file_path, sheet_name='GridFreq')
            for index, row in df_gridfreq.iterrows():
                date = row['Date']
            
                # Check if the record exists, and create it if not
                instance, created = GridFreq.objects.get_or_create(Date=date)
                # Update the fields
                instance.FreqMorning = row['MorningFreq']
                instance.FreqEvening = row['EveningFreq']
                instance.TimeMaxDemandMorning = row['MoringMaxTime']
                instance.TimeMaxDemandEvening = row['EveningMaxTime']
                instance.save()
            print('Uploaded Grid Frequency data')

        if 'SchDrwlData' in openpyxl.load_workbook(file_path).sheetnames:
            df_schdrwldata = pd.read_excel(file_path, sheet_name='SchDrwlData')
            for index, row in df_schdrwldata.iterrows():
                state_id = row['StateID']
                date = row['Date']
                if row['Exclude']==True:
                    continue
                # Try to get the existing record
                instance, created = SchDrwlData.objects.get_or_create(StateID=state_id, Date=date)
                # Update the fields
                instance.StateName = row['StateName']
                instance.Schedule = row['Schedule']
                instance.Drawl = row['Drawl']
                instance.save()
            print('Uploaded Schedule Drawl data')

        if 'CentralSectorData' in openpyxl.load_workbook(file_path).sheetnames:
            df_centralsectordata = pd.read_excel(file_path, sheet_name='CentralSectorData')
            for index, row in df_centralsectordata.iterrows():
                central_station_id = row['CentralStationID']
                date = row['Date']
                if row['Exclude']==True:
                    continue
                # Check if the record exists, and create it if not
                instance, created = CentralSectorData.objects.get_or_create(CentralStationID=central_station_id, Date=date)
                # Update the fields
                instance.CenStationName = row['CenStationName']
                instance.Energy = row['Energy']
                instance.save()
            print('Uploaded Central Sector data')

        if 'LevelStorageData' in openpyxl.load_workbook(file_path).sheetnames:
            df_levelstoragedata = pd.read_excel(file_path, sheet_name='LevelStorageData')
            for index, row in df_levelstoragedata.iterrows():
                dam_id = row['DamID']
                date = row['Date']
                # Check if the record exists, and create it if not
                instance, created = LevelStorageData.objects.get_or_create(DamID=dam_id, Date=date)
                # Update the fields
                instance.DamName = row['DamName']
                instance.Level = row['Level']
                instance.save()
            print('Uploaded Level Storage data')

        if 'InflowDischarge' in openpyxl.load_workbook(file_path).sheetnames:
            df_inflowdischarge = pd.read_excel(file_path, sheet_name='InflowDischarge')
            for index, row in df_inflowdischarge.iterrows():
                reservoir_id = row['ReservoirID']
                date = row['Date']
                # Check if the record exists, and create it if not
                instance, created = InflowsDischarge.objects.get_or_create(ReservoirID=reservoir_id, Date=date)
                # Update the fields
                instance.Type = row['Type']
                instance.Name = row['Name']
                instance.DamName = row['DamName']
                instance.InfDis00to24 = row['InfDis00to24']
                instance.InfDis06to06 = row['InfDis06to06']
                instance.save()
            print('Uploaded Inflow Discharge data')


        if 'WeatherandOtherParameters' in openpyxl.load_workbook(file_path).sheetnames:
            df_weather = pd.read_excel(file_path, sheet_name='WeatherandOtherParameters')
            for index, row in df_weather.iterrows():
                weather_id = row['WID']
                date = row['Date']
                # Check if the record exists, and create it if not
                instance, created = WeatherandOtherParameters.objects.get_or_create(WID=weather_id, Date=date)
                # Update the fields
                instance.Type = row['Type']
                instance.Name = row['Name']
                instance.Value = row['Value']
                instance.save()
            print('Uploaded Weather data')

        if 'CoalParticulars' in openpyxl.load_workbook(file_path).sheetnames:
            df_coalparticulars = pd.read_excel(file_path, sheet_name='CoalParticulars')
            for index, row in df_coalparticulars.iterrows():
                genstation_id = row['GenStationID']
                date = row['Date']
                # Check if the record exists, and create it if not
                instance, created = CoalParticulars.objects.get_or_create(GenStationID=genstation_id, Date=date)
                # Update the fields
                instance.GenStationName = row['GenStationName']
                instance.OpenBal = row['OpenBal']
                instance.Receipts = row['Receipts']
                instance.Consumption = row['Consumption']
                instance.AvgCoalperDay = row['AvgCoalperDay']
                instance.save()
            print('Uploaded Coal data')


        if 'WagonParticulars' in openpyxl.load_workbook(file_path).sheetnames:
            df_wagonparticulars = pd.read_excel(file_path, sheet_name='WagonParticulars')
            for index, row in df_wagonparticulars.iterrows():
                genstation_id = row['GenStationID']
                date = row['Date']
                # Check if the record exists, and create it if not
                instance, created = WagonParticulars.objects.get_or_create(GenStationID=genstation_id, Date=date)
                # Update the fields
                instance.GenStationName = row['GenStationName']
                instance.OpenBal = row['OpenBal']
                instance.Receipts = row['Receipts']
                instance.Tippled = row['Tippled']
                instance.Pending = row['Pending']
                instance.save()
            print('Uploaded Wagon data')

        return True, None  # Success
    except Exception as e:
        return False, str(e)  # Error message

def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            success, message = import_data_from_excel(file)
            if success:
                return render(request, 'dailyreport/success.html')
            else:
                return render(request, 'dailyreport/error.html', {'error_message': message})
    else:
        form = ExcelUploadForm()

    return render(request, 'dailyreport/upload_excel.html', {'form': form})


def user_logout(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def home(request):
    return render(request, 'dailyreport/home.html')


@login_required(login_url='login')
def export_to_text_fir(request):
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename="FIR.txt"'
    yesterday = datetime.datetime.strptime(request.GET['date'], '%Y-%m-%d').date()
    today = yesterday + relativedelta(days=1)
    yesterday_str = f'{yesterday.year}-{yesterday.month:02}-{yesterday.day:02}'
    daybeforeyesterday = today - relativedelta(days=2)
    cur_month = yesterday.month
    cur_year = yesterday.year
    cur_yearmonth = f'{cur_year}-{cur_month:02}'

    monthstartday = f'{cur_year}-{cur_month:02}-01'
    fin_year_startday = f'{cur_year-1}-04-01'

    previous_year_day = yesterday - relativedelta(years=1)
    

    query_Gen = """WITH present_data AS
               (SELECT GenStationID, GenStationName, GenType, InstalledCap, MorningPeak, EveningPeak, Energy 
               FROM dailyreport_DemandData
               WHERE Date=%(yesterday)s),
               last_year_data AS
               (SELECT GenStationID, Energy 
               FROM dailyreport_DemandData
               WHERE Date=%(previous_year_day)s)
               SELECT GenStationName, GenType, InstalledCap, p.MorningPeak, p.EveningPeak, p.Energy ,l.Energy, p.GenStationID
               FROM present_data AS p 
               LEFT OUTER JOIN last_year_data AS l
               ON p.GenStationID=l.GenStationID
               """

    query_PrevTSDemand = """
               SELECT MAX(MorningPeak, EveningPeak) 
               FROM dailyreport_DemandData
               WHERE Date=%(previous_year_day)s AND GenStationID=36
               """


    query_PrevLoadFactor = """
               SELECT Energy 
               FROM dailyreport_DemandData
               WHERE Date=%(previous_year_day)s AND GenStationID=39
               """


    query_gridfreq = """SELECT FreqMorning,FreqEvening,TimeMaxDemandMorning,TimeMaxDemandEvening
               FROM dailyreport_GridFreq
               WHERE Date=%(yesterday)s"""



    query_weatherandotherparameters = """
               SELECT WID, Name, Type, Value, Date 
               FROM dailyreport_WeatherandOtherParameters
               WHERE Date=%(yesterday)s
               """

    query_coalparticulars = """
               SELECT GenStationID, GenStationName, OpenBal, Receipts, Consumption, AvgCoalperDay, Date 
               FROM dailyreport_CoalParticulars
               WHERE Date=%(yesterday)s
               """

    query_wagonparticulars = """
               SELECT GenStationID, GenStationName, OpenBal, Receipts, Tippled, Pending, Date 
               FROM dailyreport_WagonParticulars
               WHERE Date=%(yesterday)s
               """

    query_month_gendata = f"""WITH month_data AS
               (SELECT GenStationID, GenStationName, GenType, InstalledCap, Energy, STRFTIME('%%Y-%%m', Date) AS month, Date
               FROM dailyreport_DemandData)
               
               SELECT GenStationID, GenStationName, GenType, InstalledCap, Energy, Date
               FROM month_data
               WHERE month='{cur_year}-{cur_month:02}'
               """

    query_maxcitysolar = f"""
               SELECT PID, Name, MaxDemand, Time, Date
               FROM dailyreport_maxcitysolar
               WHERE Date='{yesterday}'
               """

                
    with connection.cursor() as cursor:
        cursor.execute(query_Gen, {'yesterday': yesterday, 'previous_year_day': previous_year_day})
        gen_data = cursor.fetchall()

        cursor.execute(query_PrevTSDemand, {'previous_year_day': previous_year_day})
        PrevTSDemand = cursor.fetchall()




        cursor.execute(query_gridfreq, {'yesterday': yesterday})
        gridfreq_data = cursor.fetchall()

        cursor.execute(query_weatherandotherparameters,{'yesterday': yesterday})
        weatherandotherdata = cursor.fetchall()

        cursor.execute(query_coalparticulars,{'yesterday': yesterday})
        coaldata = cursor.fetchall()

        cursor.execute(query_wagonparticulars,{'yesterday': yesterday})
        wagondata = cursor.fetchall()

        cursor.execute(query_maxcitysolar,{'yesterday': yesterday})
        maxcitysolardata = cursor.fetchall()

        cursor.close()

        gen_data=pd.DataFrame(gen_data,columns=['GenStationName', 'GenType', 'InstalledCap', 'MorningPeak', 'EveningPeak', 'Energy' ,'PrevEnergy','GenStationID'])
        #print(gen_data)
        gridfreq_data=pd.DataFrame(gridfreq_data,columns=['FreqMorning','FreqEvening','TimeMaxDemandMorning','TimeMaxDemandEvening'])

        weatherandotherdata = pd.DataFrame(weatherandotherdata,columns=['WID', 'Name', 'Type', 'Value', 'Date'])
        weatherdata=weatherandotherdata[weatherandotherdata['Type']=='Weather']
        otherdata=weatherandotherdata[weatherandotherdata['Type']=='River']

        coaldata = pd.DataFrame(coaldata,columns=['GenStationID', 'GenStationName', 'OpenBal', 'Receipts', 'Consumption', 'AvgCoalperDay', 'Date'])
        coaldata['Balance'] = coaldata['OpenBal']+coaldata['Receipts']-coaldata['Consumption']

        wagondata = pd.DataFrame(wagondata,columns=['GenStationID', 'GenStationName', 'OpenBal', 'Receipts', 'Tippled', 'Pending', 'Date'])

        maxcitysolardata = pd.DataFrame(maxcitysolardata,columns=['PID', 'Name', 'MaxDemand', 'Time','Date'])



        thermal=gen_data[gen_data["GenType"] == 'Thermal']

        hydel=gen_data[gen_data["GenType"] == 'Hydel']

        genco=pd.concat([thermal,hydel],axis=0)

        genco_total=genco.sum()

        central_sector=gen_data[gen_data["GenType"] == 'Central Sector']

        lta=gen_data[gen_data["GenType"] == 'LTA']
        
        APISGS=gen_data[gen_data["GenType"] == 'APISGS']

        solar=gen_data[gen_data["GenType"] == 'Private_solar']

        wind=gen_data[gen_data["GenType"] == 'Private_wind']

        nonconventional=gen_data[gen_data["GenType"] == 'Private_Nonconventional']

        private=gen_data[gen_data["GenType"] == 'Private']

        private_total=pd.concat([private,wind,solar,nonconventional],axis=0)

        state_purchases=gen_data[gen_data["GenType"] == 'State Purchases']

        third_party_purchase=gen_data[gen_data["GenType"] == 'Third Party Purchases']

        third_party_sales=gen_data[gen_data["GenType"] == 'Third Party Sales']

        gen_total=pd.concat([genco,central_sector,lta,APISGS,private_total,state_purchases,third_party_purchase,third_party_sales],axis=0)
        gen_total=gen_total[['InstalledCap','MorningPeak','EveningPeak','Energy','PrevEnergy']].sum()

        pump=gen_data[gen_data["GenType"] == 'Pump']

        pump_total=pump[['InstalledCap','MorningPeak','EveningPeak','Energy','PrevEnergy']].sum()

        gen_total_wo_pump=gen_total.add(pump_total)

        load_factor=(gen_total_wo_pump["Energy"]*1000*100/24/max(gen_total_wo_pump["MorningPeak"],gen_total_wo_pump["EveningPeak"])).round(2)

        instance, created = DemandData.objects.get_or_create(GenStationID=39, Date=yesterday)
        instance.Energy = load_factor
        instance.GenType = 'Total'
        instance.save()

    with connection.cursor() as cursor:
        cursor.execute(query_PrevLoadFactor, {'previous_year_day': previous_year_day})
        PrevLoadFactor = cursor.fetchall()

        cursor.close()
        
    subindex=['i','ii','iii','iv','v','vi','vii','viii']
    alphabets=['a','b','c','d','e','f','g']

    
    

    # Create the report content as a string
    report_content = f"""
                 TRANSMISSION CORPORATION OF TELANGANA LTD
               GRID OPERATION -- INITIAL REPORT FOR {yesterday.strftime('%d/%m/%Y')}
=================================================================================
   |     Generating Station     |Generation at Peak Demand|     Generation  
   |                            |      EX-BUS in MW       |    EX-BUS in MU
Sl |                            |  Morning      Evening   | {yesterday.strftime('%A'):^10}|{previous_year_day.strftime('%A'):^10}
No |                            |  {gridfreq_data.iloc[0,0]:.2f}HZ      {gridfreq_data.iloc[0,1]:.2f}HZ   | {yesterday.strftime('%d/%m/%Y')}|{previous_year_day.strftime('%d/%m/%Y')}
   |   Name            INS.CAP  |  {gridfreq_data.iloc[0,2].strftime('%H:%M')}Hrs     {gridfreq_data.iloc[0,3].strftime('%H:%M')}Hrs  |           |
-----------------------(MW)------------------------------------------------------
(1)     (2)                         (3)         (4)           (5)           (6)
---------------------------------------------------------------------------------
 I  TS GENCO"""

    # Add the data rows to the report content
    for i in range(hydel.shape[0]):
      row_content = f"""
     {hydel.iloc[i,0]:<17}{hydel.iloc[i,2]:>6.0f}{hydel.iloc[i,3]:>12.0f}{hydel.iloc[i,4]:>12.0f}{hydel.iloc[i,5]:>14.3f}    |{hydel.iloc[i,6]:>10.3f}"""
      report_content += row_content

    row_content = f"""
       TS HYDEL-->    {hydel["InstalledCap"].sum():>6.0f}{hydel["MorningPeak"].sum():>12.0f}{hydel["EveningPeak"].sum():>12.0f}{hydel["Energy"].sum():>14.3f}    |{hydel["PrevEnergy"].sum():>10.3f}
        """       
    report_content += row_content


    for i in range(thermal.shape[0]):
      row_content = f"""
     {thermal.iloc[i,0]:<17}{thermal.iloc[i,2]:>6.0f}{thermal.iloc[i,3]:>12.0f}{thermal.iloc[i,4]:>12.0f}{thermal.iloc[i,5]:>14.3f}    |{thermal.iloc[i,6]:>10.3f}"""
      report_content += row_content

    row_content = f"""
       TS THERMAL-->  {thermal["InstalledCap"].sum():>6.0f}{thermal["MorningPeak"].sum():>12.0f}{thermal["EveningPeak"].sum():>12.0f}{thermal["Energy"].sum():>14.3f}    |{thermal["PrevEnergy"].sum():>10.3f}"""       
    report_content += row_content

    row_content = f"""

       TSGENCO TOTAL->{genco["InstalledCap"].sum():>6.0f}{genco["MorningPeak"].sum():>12.0f}{genco["EveningPeak"].sum():>12.0f}{genco["Energy"].sum():>14.3f}    |{genco["PrevEnergy"].sum():>10.3f}"""       
    report_content += row_content

    for i in range(lta.shape[0]):
      row_content = f"""

     {lta.iloc[i,0]:<17}{lta.iloc[i,2]:>6.0f}{lta.iloc[i,3]:>12.0f}{lta.iloc[i,4]:>12.0f}{lta.iloc[i,5]:>14.3f}    |{lta.iloc[i,6]:>10.3f}"""
      report_content += row_content

    row_content = f"""

II CENTRAL SECTOR"""       
    report_content += row_content
    row_content = f"""
     {central_sector.iloc[0,0]:<17}{central_sector.iloc[0,2]:>6.0f}{central_sector.iloc[0,3]:>12.0f}{central_sector.iloc[0,4]:>12.0f}{central_sector.iloc[0,5]:>14.3f}    |{central_sector.iloc[0,6]:>10.3f}"""
    report_content += row_content

    row_content = f"""

III TSShare OF APISGS->{APISGS.iloc[0,2]:>5.0f}{APISGS.iloc[0,3]:>12.0f}{APISGS.iloc[0,4]:>12.0f}{APISGS.iloc[0,5].round(2):>14.3f}    |{APISGS.iloc[0,6]:>10.3f}"""       
    report_content += row_content

    row_content = f"""


                 TRANSMISSION CORPORATION OF TELANGANA LTD
               GRID OPERATION -- INITIAL REPORT FOR {yesterday.strftime('%d/%m/%Y')}
==================================================================================
   |     Generating Station     |Generation at Peak Demand|        Generation  
   |                            |      EX-BUS in MW       |       EX-BUS in MU
Sl |                            |  Morning      Evening   |{yesterday.strftime('%A'):^12}|{previous_year_day.strftime('%A'):^10}
No |                            |  {gridfreq_data.iloc[0,0]:.2f}HZ      {gridfreq_data.iloc[0,1]:.2f}HZ   |{yesterday.strftime('%d/%m/%Y'):^12}|{previous_year_day.strftime('%d/%m/%Y'):^10}
   |   Name            INS.CAP  |  {gridfreq_data.iloc[0,2].strftime('%H:%M')}Hrs     {gridfreq_data.iloc[0,3].strftime('%H:%M')}Hrs  |            |
-----------------------(MW)-------------------------------------------------------
(1)    (2)                          (3)          (4)            (5)         (6)
----------------------------------------------------------------------------------
"""
    report_content += row_content

    row_content = f"""
IV  PRIVATE SECTOR

     a) """   
#    print(private)
    report_content += row_content
    row_content = f"""{private.iloc[0,0]:<14}{private.iloc[0,2]:>6.0f}{private.iloc[0,3]:>12.0f}{private.iloc[0,4]:>12.0f}{private.iloc[0,5]:>16.3f}   |{private.iloc[0,6]:>8.3f}"""
    report_content += row_content
#    row_content = f"""{private.iloc[1,0]:<14}{private.iloc[1,2]:>6.1f}{private.iloc[1,3]:>10.0f}{private.iloc[1,4]:>12.0f}{private.iloc[1,5]:>16.3f}     |{private.iloc[1,6]:>8.3f}"""
#    report_content += row_content
    row_content = f"""
     b) """       
    report_content += row_content
    for i in range(wind.shape[0]):
      row_content = f"""{wind.iloc[i,0]:<14}{wind.iloc[i,2]:>6.1f}{wind.iloc[i,3]:>12.0f}{wind.iloc[i,4]:>12.0f}{wind.iloc[i,5]:>16.3f}   |{wind.iloc[i,6]:>8.3f}"""
      report_content += row_content

    row_content = f"""

     c) TOTAL SOLAR      {solar["InstalledCap"].sum():>6.2f}"""       
    report_content += row_content
    for i in range(solar.shape[0]):
      row_content = f"""
       {subindex[i]}){' '*(1-i)} {solar.iloc[i,0][:6]:<11}{solar.iloc[i,2]:>6.2f}{solar.iloc[i,3]:>11.0f}{solar.iloc[i,4]:>12.0f}{solar.iloc[i,5]:>16.3f}   |{solar.iloc[i,6]:>8.3f}"""
      report_content += row_content
    #print(solar)
    row_content = f"""

     d) NONCONVENTIONAL   {nonconventional["InstalledCap"].sum():>5.1f}{' ':12}{'':12}"""       
    report_content += row_content
    for i in range(nonconventional.shape[0]):
        if i==0:
            row_content = f"""
       {subindex[i]}){'  '}{nonconventional.iloc[i,0]:<18}{nonconventional.iloc[i,3]:>11.0f}{nonconventional.iloc[i,4]:>12.0f}{nonconventional.iloc[i,5]:>16.3f}   |{nonconventional.iloc[i,6]:>8.3f}"""
            report_content += row_content
        else:
            row_content = f"""
       {subindex[i]}){' '}{nonconventional.iloc[i,0]:<18}{'':>11}{'':>12}{nonconventional.iloc[i,5]:>16.3f}   |{nonconventional.iloc[i,6]:>8.3f}"""
            report_content += row_content
 
    row_content = f"""

     PRIVATE SECTOR TOTAL   {private_total["MorningPeak"].sum():12.0f}{private_total["EveningPeak"].sum():12.0f}{private_total["Energy"].sum():>16.3f}   |{private_total["PrevEnergy"].sum():>8.3f}"""       
    report_content += row_content

    row_content = f"""

V   STATE PURCHASES   {'':>6}{' ':12}{' ':12}"""       
    report_content += row_content

    for i in range(state_purchases.shape[0]):
      row_content = f"""
     {alphabets[i]}) {state_purchases.iloc[i,0]:<14}{'':>6}{state_purchases.iloc[i,3]:>12.0f}{state_purchases.iloc[i,4]:>12.0f}{state_purchases.iloc[i,5]:>16.3f}   |{state_purchases.iloc[i,6]:>8.3f}"""
      report_content += row_content
    row_content = f"""

     STATE PURCHASE TOTAL   {state_purchases["MorningPeak"].sum():12.0f}{state_purchases["EveningPeak"].sum():12.0f}{state_purchases["Energy"].sum():>16.3f}   |{state_purchases["PrevEnergy"].sum():>8.3f}"""
    report_content += row_content
    row_content = f"""

VI  THIRD PARTY PURCHASES   {third_party_purchase.iloc[0,3]:>12.0f}{third_party_purchase.iloc[0,4]:>12.0f}{third_party_purchase.iloc[0,5]:>16.3f}   |{third_party_purchase.iloc[0,6]:>8.3f}
    (CONSUMER PURCHASES)"""
    report_content += row_content

    row_content = f"""

VII THIRD PARTY SALES       {third_party_sales.iloc[0,3]:>12.0f}{third_party_sales.iloc[0,4]:>12.0f}{third_party_sales.iloc[0,5]:>16.3f}   |{third_party_sales.iloc[0,6]:>8.3f}
    (INTRA STATE PVT GENERATORS)"""
    report_content += row_content

    row_content = f"""

VIII TOTAL DEMAND & CONSUMP {gen_total["MorningPeak"]:>12.0f}{gen_total["EveningPeak"]:>12.0f}{gen_total["Energy"]:>16.3f}   |{gen_total["PrevEnergy"]:>8.3f}
     (WITH PUMPS)"""
    report_content += row_content

    row_content = f"""

IX  {pump.iloc[0,0]:<24}{pump.iloc[0,3]:>12.0f}{pump.iloc[0,4]:>12.0f}{pump.iloc[0,5]:>16.3f}   |{pump.iloc[0,6]:>8.3f}"""
    report_content += row_content

    row_content = f"""


X   {pump.iloc[1,0]:<24}{pump.iloc[1,3]:>12.0f}{pump.iloc[1,4]:>12.0f}{pump.iloc[1,5]:>16.3f}   |{pump.iloc[1,6]:>8.3f}
    """
    report_content += row_content

    row_content = f"""

XI  TS DEMAND(EX-BUS) {gen_total_wo_pump["InstalledCap"]:<6.0f}{gen_total_wo_pump["MorningPeak"]:>12.0f}{gen_total_wo_pump["EveningPeak"]:>12.0f}{'':>16}   |{PrevTSDemand[0][0]:>8.0f}


       ENERGY (MU)    {'':>6}{'':>12}{'':>12}{gen_total_wo_pump["Energy"]:>16.3f}   |{gen_total_wo_pump["PrevEnergy"]:>8.3f}"""
    report_content += row_content

    row_content = f"""


XII  Load Factor      {'':<6}{'':>12}{'':>12}{load_factor:>16.3f}%  |{PrevLoadFactor[0][0]:>8.3f}%
"""
    report_content += row_content


    romannumerals=['XIII','XIV','XV','XVI']
    for i in range(maxcitysolardata.shape[0]):
        row_content = f"""
{romannumerals[i]:<5}{maxcitysolardata.iloc[i,1]:<25}{maxcitysolardata.iloc[i,2]:>5.0f} MW    AT   {str(maxcitysolardata.iloc[i,3])[:5]:>8} Hrs
"""   
        #print(row_content)
        report_content += row_content



    row_content = f"""



              STATUS OF COAL SUPPLIES TO THERMAL STATIONS ON :  {yesterday.strftime('%d/%m/%Y')}
  ===============================================================================
   {'Station':<10}{'Op. Balance':^12}{'Receipts':^12}{'Consumption':^12}{'Balance':^12}{'Average coal':^15} 
   {'':^10}{'(MTs)':^12}{'(MTs)':^12}{'(MTs)':^12}{'(MTs)':^12}{'required/day for':^15}
   {'':^10}{'':^12}{'':^12}{'':^12}{'':^12}{'full Generation(MTs)':^15}
  -------------------------------------------------------------------------------
"""           
    report_content += row_content


    for i in range(coaldata.shape[0]):
      if coaldata.iloc[i,0]==4:
          row_content = f"""  {coaldata.iloc[i,1]:<10}{coaldata.iloc[i,2]:>10}{coaldata.iloc[i,3]:>12}{coaldata.iloc[i,4]:>12}{coaldata.iloc[i,7]-coaldata.iloc[i+1,4]:>12}{coaldata.iloc[i,5]:>15}
"""
      elif coaldata.iloc[i,0]==5:
          row_content = f"""  {coaldata.iloc[i,1]:<10}{'':<10}{'':>12}{coaldata.iloc[i,4]:>12}{'':>12}{coaldata.iloc[i,5]:>15}
"""
      else:
          row_content = f"""  {coaldata.iloc[i,1]:<10}{coaldata.iloc[i,2]:>10}{coaldata.iloc[i,3]:>12}{coaldata.iloc[i,4]:12}{coaldata.iloc[i,7]:>12}{coaldata.iloc[i,5]:>15}
"""
      report_content += row_content
    row_content = f"""  ===============================================================================
"""
    report_content += row_content


    row_content = f"""

                                COAL WAGONS POSITION
  ===============================================================================
   {'Station':<15}{'Op. Balance':^17}{'Receipts':^17}{'Consumption':^17}{'Balance':^17}
  -------------------------------------------------------------------------------
"""           
    report_content += row_content

    for i in range(wagondata.shape[0]):
      row_content = f"""  {wagondata.iloc[i,1]:<12}{wagondata.iloc[i,2]:>13}{wagondata.iloc[i,3]:>17}{wagondata.iloc[i,4]:>17}{wagondata.iloc[i,5]:>17}
"""
      report_content += row_content
    row_content = f"""  ===============================================================================
"""
    report_content += row_content
    if gen_total_wo_pump["MorningPeak"]>gen_total_wo_pump["EveningPeak"]:
        row_content=f"""
        
        
        
    
        
Sir,
Energy Demand Particulars for Dt: {yesterday.strftime('%d.%m.%Y')}
Hydel- {hydel["MorningPeak"].sum():.0f} MW, {hydel["Energy"].sum():.3f} MU;
Thermal- {thermal["MorningPeak"].sum():.0f} MW, {thermal["Energy"].sum():.3f} MU;
Singareni- {lta.loc[lta['GenStationID']==18].MorningPeak.values[0]:.0f} MW, {lta.loc[lta['GenStationID']==18].Energy.values[0]:.3f} MU;
Chattisgarh- {lta.loc[lta['GenStationID']==20].MorningPeak.values[0]:.0f} MW, {lta.loc[lta['GenStationID']==20].Energy.values[0]:.3f} MU;
CGS- {central_sector.loc[central_sector['GenStationID']==21].MorningPeak.values[0]:.0f} MW, {central_sector.loc[central_sector['GenStationID']==21].Energy.values[0]:.3f} MU;
TS Share of APISGS- {APISGS.loc[APISGS['GenStationID']==22].MorningPeak.values[0]:.0f} MW, {APISGS.loc[APISGS['GenStationID']==22].Energy.values[0]:.3f} MU;
Pvt Sector- {private_total["MorningPeak"].sum():.0f} MW, {private_total["Energy"].sum():.3f} MU;
State Purchases- {state_purchases["MorningPeak"].sum():.0f} MW, {state_purchases["Energy"].sum():.3f} MU;
Third Party Purchases- {third_party_purchase["MorningPeak"].sum():.0f} MW, {third_party_purchase["Energy"].sum():.3f} MU;
Third Party Sales- {third_party_sales["MorningPeak"].sum():.0f} MW, {third_party_sales["Energy"].sum():.3f} MU;
SSLM PUMP Consumption- {pump.iloc[0,3]:.0f} MW, {pump.iloc[0,5]:.3f} MU;
NSR PUMP Consumption- {pump.iloc[1,3]:.0f} MW, {pump.iloc[1,5]:.3f} MU;

Total consumption: {gen_total_wo_pump["Energy"]:.3f} MU;

Max. Demand met: {gen_total["MorningPeak"]:.0f} MW @ {gridfreq_data.iloc[0,2].strftime('%H:%M')}Hrs (Including Solar & NSR Pumps)

Min. Demand met: {maxcitysolardata.iloc[2,2]:.0f} MW @ {str(maxcitysolardata.iloc[2,3])[:5]} Hrs.

Max. Solar: {maxcitysolardata.iloc[1,2]:.0f} MW @ {str(maxcitysolardata.iloc[1,3])[:5]} Hrs.

Solar Energy: {solar["Energy"].sum():.3f} MU.

Max. City Demand: {maxcitysolardata.iloc[0,2]:.0f} MW @ {str(maxcitysolardata.iloc[0,3])[:5]} Hrs.

Max. City temperature:       C.


        """


    else:
        row_content=f"""
        
        
Sir,
Energy Demand Particulars for Dt: {yesterday.strftime('%d.%m.%Y')}
Hydel- {hydel["EveningPeak"].sum():.0f} MW, {hydel["Energy"].sum():.3f} MU;
Thermal- {thermal["EveningPeak"].sum():.0f} MW, {thermal["Energy"].sum():.3f} MU;
Singareni- {lta.loc[lta['GenStationID']==18].EveningPeak.values[0]:.0f} MW, {lta.loc[lta['GenStationID']==18].Energy.values[0]:.3f} MU;
Chattisgarh- {lta.loc[lta['GenStationID']==20].EveningPeak.values[0]:.0f} MW, {lta.loc[lta['GenStationID']==20].Energy.values[0]:.3f} MU;
CGS- {central_sector.loc[central_sector['GenStationID']==21].EveningPeak.values[0]:.0f} MW, {central_sector.loc[central_sector['GenStationID']==21].Energy.values[0]:.3f} MU;
TS Share of APISGS- {APISGS.loc[APISGS['GenStationID']==22].EveningPeak.values[0]:.0f} MW, {APISGS.loc[APISGS['GenStationID']==22].Energy.values[0]:.3f} MU;
Pvt Sector-{private_total["EveningPeak"].sum():.0f} MW, {private_total["Energy"].sum():.3f} MU;
State Purchases-{state_purchases["EveningPeak"].sum():.0f} MW, {state_purchases["Energy"].sum():.3f} MU;
Third Party Purchases-{third_party_purchase["EveningPeak"].sum():.0f} MW, {third_party_purchase["Energy"].sum():.3f} MU;
Third Party Sales-{third_party_sales["EveningPeak"].sum():.0f} MW, {third_party_sales["Energy"].sum():.3f} MU;
SSLM PUMP Consumption- {pump.iloc[0,4]:.0f} MW, {pump.iloc[0,5]:.3f} MU;
NSR PUMP Consumption- {pump.iloc[1,4]:.0f} MW, {pump.iloc[1,5]:.3f} MU;

Total consumption: {gen_total_wo_pump["Energy"]:.3f} MU;

Max. Demand met: {gen_total["EveningPeak"]:.0f} MW @ {gridfreq_data.iloc[0,3].strftime('%H:%M')}Hrs (Including Solar & NSR Pumps)

Min. Demand met: {maxcitysolardata.iloc[2,2]:.0f} MW @ {str(maxcitysolardata.iloc[2,3])[:5]} Hrs.

Max. Solar: {maxcitysolardata.iloc[1,2]:.0f} MW @ {str(maxcitysolardata.iloc[1,3])[:5]} Hrs.

Solar Energy: {solar["Energy"].sum()} MU.

Max. City Demand: {maxcitysolardata.iloc[0,2]:.0f} MW @ {str(maxcitysolardata.iloc[0,3])[:5]} Hrs.

Max. City temperature:       C.

        """       
        
    report_content += row_content
    response.write(report_content)

    return response


@login_required(login_url='login')
def export_to_text(request):
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename="RPT.txt"'
    yesterday = datetime.datetime.strptime(request.GET['date'], '%Y-%m-%d').date()
    today = yesterday + relativedelta(days=1)
    yesterday_str = f'{yesterday.year}-{yesterday.month:02}-{yesterday.day:02}'
    daybeforeyesterday = today - relativedelta(days=2)
    cur_month = f'{yesterday.month:02}'
    cur_year = yesterday.year
    cur_yearmonth = f'{cur_year}-{cur_month:02}'

    monthstartday = f'{cur_year}-{cur_month:02}-01'

    if int(cur_month)<4:
        fin_year_startday = f'{cur_year-1}-04-01'
    else:
        fin_year_startday = f'{cur_year}-04-01'
    #print(monthstartday)
    previous_year_day = yesterday - relativedelta(years=1)
    previous_year_today = today - relativedelta(years=1)
    

    query_Gen = """WITH present_data AS
               (SELECT GenStationID, GenStationName, GenType, InstalledCap, MorningPeak, EveningPeak, Energy 
               FROM dailyreport_DemandData
               WHERE Date=%(yesterday)s),
               last_year_data AS
               (SELECT GenStationID, Energy 
               FROM dailyreport_DemandData
               WHERE Date=%(previous_year_day)s)
               SELECT GenStationName, GenType, InstalledCap, p.MorningPeak, p.EveningPeak, p.Energy ,l.Energy, p.GenStationID
               FROM present_data AS p 
               LEFT OUTER JOIN last_year_data AS l
               ON p.GenStationID=l.GenStationID
               """

    query_PrevTSDemand = """
               SELECT MAX(MorningPeak, EveningPeak) 
               FROM dailyreport_DemandData
               WHERE Date=%(previous_year_day)s AND GenStationID=36
               """


    query_PrevLoadFactor = """
               SELECT Energy 
               FROM dailyreport_DemandData
               WHERE Date=%(previous_year_day)s AND GenStationID=39
               """


#    query_monthdataTSDemand = """
#               SELECT MorningPeak, EveningPeak, Energy, Date
#              FROM dailyreport_DemandData
#               WHERE (Date BETWEEN %(monthstartday)s AND %(yesterday)s) AND GenStationID=36
#              """


    query_TsdemandMonthCum = """
               SELECT SUM(Energy) 
               FROM dailyreport_DemandData
               WHERE (Date BETWEEN %(monthstartday)s AND %(yesterday)s) AND GenStationID=36
               """

    query_TsdemandMonthCumCorrection = """
               SELECT SUM(Energy) 
               FROM dailyreport_DemandData
               WHERE (Date BETWEEN %(monthstartday)s AND %(yesterday)s) AND GenStationID=39
               """


    query_TsdemandFinYearCum = """
               SELECT SUM(Energy) 
               FROM dailyreport_DemandData
               WHERE (Date BETWEEN %(fin_year_startday)s AND %(yesterday)s) AND GenStationID=36
               """

    query_TsdemandFinYearCumCorrection = """
               SELECT SUM(Energy) 
               FROM dailyreport_DemandData
               WHERE (Date BETWEEN %(fin_year_startday)s AND %(yesterday)s) AND GenStationID=39
               """

    query_gridfreq = """SELECT FreqMorning,FreqEvening,TimeMaxDemandMorning,TimeMaxDemandEvening
               FROM dailyreport_GridFreq
               WHERE Date=%(yesterday)s"""

    query_centralgen = f"""
               SELECT CentralStationID, CenStationName, Energy, Date
               FROM dailyreport_centralsectordata
               WHERE Date BETWEEN %(monthstartday)s AND %(yesterday)s
               """

    query_schdrwl = f"""
               SELECT StateID, StateName, Schedule, Drawl, Date
               FROM dailyreport_SchDrwlData
               WHERE Date BETWEEN %(monthstartday)s AND %(yesterday)s
               """

    query_levelstorage = f"""
               SELECT *
               FROM dailyreport_LevelStorage
               """
    query_levelstoragedata = """
               WITH present_data AS
               (SELECT d.DamID, d.DamName, d.Level, ls.Storage 
               FROM dailyreport_LevelStorageData AS d
               LEFT OUTER JOIN dailyreport_LevelStorage AS ls
               ON d.DamID=ls.DamID AND d.Level=ls.Level
               WHERE Date=%(today)s),

               last_year_data AS
               (SELECT d.DamID, d.DamName, d.Level, ls.Storage 
               FROM dailyreport_LevelStorageData AS d
               LEFT OUTER JOIN dailyreport_LevelStorage AS ls
               ON d.DamID=ls.DamID AND d.Level=ls.Level
               WHERE Date=%(previous_year_today)s),

               yesterday_data AS
               (SELECT DamID, Level
               FROM dailyreport_LevelStorageData AS d
               WHERE Date=%(yesterday)s)

               SELECT p.DamID, p.DamName, l.Level, l.Storage, p.Level, p.Storage, p.Level-y.Level AS LevelRaise
               FROM present_data AS p 
               LEFT OUTER JOIN last_year_data AS l
               ON p.DamID=l.DamID
               LEFT OUTER JOIN yesterday_data AS y
               ON p.DamID=y.DamID
               """

    query_inflowsdischarge = """
               SELECT ReservoirID, Name, Type, DamName, InfDis00to24, InfDis06to06, Date 
               FROM dailyreport_InflowsDischarge
               WHERE Date=%(yesterday)s
               """

    query_weatherandotherparameters = """
               SELECT WID, Name, Type, Value, Date 
               FROM dailyreport_WeatherandOtherParameters
               WHERE Date=%(yesterday)s
               """

    query_coalparticulars = """
               SELECT GenStationID, GenStationName, OpenBal, Receipts, Consumption, AvgCoalperDay, Date 
               FROM dailyreport_CoalParticulars
               WHERE Date=%(yesterday)s
               """

    query_wagonparticulars = """
               SELECT GenStationID, GenStationName, OpenBal, Receipts, Tippled, Pending, Date 
               FROM dailyreport_WagonParticulars
               WHERE Date=%(yesterday)s
               """

    query_month_gendata = f"""WITH month_data AS
               (SELECT GenStationID, GenStationName, GenType, InstalledCap, Energy, STRFTIME('%%Y-%%m', Date) AS month, Date
               FROM dailyreport_DemandData)
               
               SELECT GenStationID, GenStationName, GenType, InstalledCap, Energy, Date
               FROM month_data
               WHERE month='{cur_year}-{cur_month:02}'
               """

    query_month_maxcitysolar = f"""
               WITH month_data AS
               (SELECT PID, Name, MaxDemand, Time, STRFTIME('%%Y-%%m', Date) AS month, Date
               FROM dailyreport_MaxCitySolar)
               
               SELECT PID, Name, MaxDemand, Time, Date
               FROM month_data
               WHERE month='{cur_year}-{cur_month:02}'
               """

                
    with connection.cursor() as cursor:
        cursor.execute(query_Gen, {'yesterday': yesterday, 'previous_year_day': previous_year_day})
        gen_data = cursor.fetchall()

        cursor.execute(query_PrevTSDemand, {'previous_year_day': previous_year_day})
        PrevTSDemand = cursor.fetchall()

#        cursor.execute(query_monthdataTSDemand, {'yesterday': yesterday, 'monthstartday': monthstartday})
#        tsdemand_monthdata = cursor.fetchall()

        cursor.execute(query_TsdemandMonthCum, {'yesterday': yesterday, 'monthstartday': monthstartday})
        tsdemand_monthcum = cursor.fetchall()

        instance, created = DemandData.objects.get_or_create(GenStationID=37, Date=yesterday)
        instance.Energy = tsdemand_monthcum[0][0]
        instance.save()

        cursor.execute(query_gridfreq, {'yesterday': yesterday})
        gridfreq_data = cursor.fetchall()

        cursor.execute(query_centralgen,{'yesterday': yesterday,  'monthstartday': monthstartday})
        centralgendata = cursor.fetchall()

        cursor.execute(query_schdrwl,{'yesterday': yesterday,  'monthstartday': monthstartday})
        schdrwldata = cursor.fetchall()

        cursor.execute(query_levelstoragedata,{'today': today, 'previous_year_today': previous_year_today, 'yesterday':yesterday})
        levelstoragedata = cursor.fetchall()

        cursor.execute(query_inflowsdischarge,{'yesterday': yesterday})
        inflowsdischargedata = cursor.fetchall()

        cursor.execute(query_weatherandotherparameters,{'yesterday': yesterday})
        weatherandotherdata = cursor.fetchall()

        cursor.execute(query_month_gendata,{'cur_year':cur_year,'cur_month':cur_month})
        monthgendata = cursor.fetchall()

        cursor.execute(query_month_maxcitysolar,{'cur_year':cur_year,'cur_month':cur_month})
        monthmaxcitysolardata = cursor.fetchall()

        cursor.close()

        gen_data=pd.DataFrame(gen_data,columns=['GenStationName', 'GenType', 'InstalledCap', 'MorningPeak', 'EveningPeak', 'Energy' ,'PrevEnergy','GenStationID'])

#        tsdemand_monthdata=pd.DataFrame(tsdemand_monthdata,columns=['MorningPeak', 'EveningPeak', 'Energy', 'Date'])
#        tsdemand_monthdata['MaxTSDemand']=tsdemand_monthdata[['MorningPeak', 'EveningPeak']].max(axis=1)

        gridfreq_data=pd.DataFrame(gridfreq_data,columns=['FreqMorning','FreqEvening','TimeMaxDemandMorning','TimeMaxDemandEvening'])

        centralgendata=pd.DataFrame(centralgendata,columns=['CentralStationID', 'CenStationName', 'Energy', 'Date'])

        schdrwldata=pd.DataFrame(schdrwldata,columns=['StateID', 'StateName', 'Schedule', 'Drawl', 'Date'])

        inflowsdischargedata=pd.DataFrame(inflowsdischargedata,columns=['ReservoirID', 'Name', 'Type', 'DamName', 'InfDis00to24', 'InfDis06to06', 'Date'])
        inf_ujurala = inflowsdischargedata[(inflowsdischargedata['Type']=='Inflow') & (inflowsdischargedata['DamName']=='Upper Jurala')]
        inf_nsagar = inflowsdischargedata[(inflowsdischargedata['Type']=='Inflow') & (inflowsdischargedata['DamName']=='NSagar')]
        inf_srisailam = inflowsdischargedata[(inflowsdischargedata['Type']=='Inflow') & (inflowsdischargedata['DamName']=='Srisailam')]
        inf_pulichintala = inflowsdischargedata[(inflowsdischargedata['Type']=='Inflow') & (inflowsdischargedata['DamName']=='Pulichintala')]
        dis_ujurala = inflowsdischargedata[(inflowsdischargedata['Type']=='Discharge') & (inflowsdischargedata['DamName']=='Upper Jurala')]
        dis_ljurala = inflowsdischargedata[(inflowsdischargedata['Type']=='Discharge') & (inflowsdischargedata['DamName']=='Lower Jurala')]
        dis_nsagar = inflowsdischargedata[(inflowsdischargedata['Type']=='Discharge') & (inflowsdischargedata['DamName']=='NSagar')]
        dis_srisailam= inflowsdischargedata[(inflowsdischargedata['Type']=='Discharge') & (inflowsdischargedata['DamName']=='Srisailam')]
        dis_pulichintala= inflowsdischargedata[(inflowsdischargedata['Type']=='Discharge') & (inflowsdischargedata['DamName']=='Pulichintala')]

        levelstoragedata=pd.DataFrame(levelstoragedata,columns=['DamID','DamName', 'PrevLevel', 'PrevStorage', 'Level', 'Storage', 'LevelRaise'])

        if levelstoragedata['Storage'].dtypes=='object':
            levelstoragedata['EquivalentEnergy']=np.nan
        else:
            levelstoragedata['EquivalentEnergy']=levelstoragedata['Storage']*np.array([2,1.279,5.5,5.5,1.513,1,1.38,1])
        levelstoragedata['FRL']=np.array([1633,1045,885,590,175,1405,1091,1718])
        levelstoragedata['MDDL']=np.array([1582,1033,800,510,140,1376,1064,1699])

        weatherandotherdata = pd.DataFrame(weatherandotherdata,columns=['WID', 'Name', 'Type', 'Value', 'Date'])
        weatherdata=weatherandotherdata[weatherandotherdata['Type']=='Weather']
        otherdata=weatherandotherdata[weatherandotherdata['Type']=='River']

        monthgendata = pd.DataFrame(monthgendata,columns=['GenStationID', 'GenStationName', 'GenType', 'InstalledCap', 'Energy', 'Date'])

        monthmaxcitysolardata = pd.DataFrame(monthmaxcitysolardata,columns=['PID','Name','MaxDemand','Time','Date'])
        maxcitysolardata = monthmaxcitysolardata[monthmaxcitysolardata['Date']==yesterday]

        centralgendata_cum=centralgendata[['CentralStationID','Energy']].groupby(['CentralStationID']).sum()
        centralgendata_cum.rename({'Energy':'MonthCumulative'},inplace=True,axis=1)
        centralgendata_today=centralgendata[centralgendata['Date']==yesterday]
        centralgendata_today=centralgendata_today.merge(centralgendata_cum,how='left',on='CentralStationID')

        schdrwldata_cum=schdrwldata[['StateID','Schedule','Drawl']].groupby(['StateID']).sum()
        schdrwldata_cum.rename({'Schedule':'MonthCumSch','Drawl':'MonthCumDrawl'},inplace=True,axis=1)
        schdrwldata_today=schdrwldata[schdrwldata.Date==yesterday]
        schdrwldata_today=schdrwldata_today.merge(schdrwldata_cum,how='left',on='StateID')
        schdrwldata_today['Diff']=schdrwldata_today['Drawl']-schdrwldata_today['Schedule']
        schdrwldata_today['CumDiff']=schdrwldata_today['MonthCumDrawl']-schdrwldata_today['MonthCumSch']
        schdrwldata_today = schdrwldata_today.sort_values(by='StateID')


        thermal=gen_data[gen_data["GenType"] == 'Thermal']

        hydel=gen_data[gen_data["GenType"] == 'Hydel']

        genco=pd.concat([thermal,hydel],axis=0)

        genco_total=genco.sum()

        central_sector=gen_data[gen_data["GenType"] == 'Central Sector']

        lta=gen_data[gen_data["GenType"] == 'LTA']
        
        APISGS=gen_data[gen_data["GenType"] == 'APISGS']

        solar=gen_data[gen_data["GenType"] == 'Private_solar']

        wind=gen_data[gen_data["GenType"] == 'Private_wind']

        nonconventional=gen_data[gen_data["GenType"] == 'Private_Nonconventional']

        private=gen_data[gen_data["GenType"] == 'Private']

        private_total=pd.concat([private,wind,solar,nonconventional],axis=0)

        state_purchases=gen_data[gen_data["GenType"] == 'State Purchases']

        third_party_purchase=gen_data[gen_data["GenType"] == 'Third Party Purchases']

        third_party_sales=gen_data[gen_data["GenType"] == 'Third Party Sales']

        gen_total=pd.concat([genco,central_sector,lta,APISGS,private_total,state_purchases,third_party_purchase,third_party_sales],axis=0)
        gen_total=gen_total[['InstalledCap','MorningPeak','EveningPeak','Energy','PrevEnergy']].sum()

        pump=gen_data[gen_data["GenType"] == 'Pump']

        pump_total=pump[['InstalledCap','MorningPeak','EveningPeak','Energy','PrevEnergy']].sum()

        gen_total_wo_pump=gen_total.add(pump_total)

        load_factor=(gen_total_wo_pump["Energy"]*1000*100/24/max(gen_total_wo_pump["MorningPeak"],gen_total_wo_pump["EveningPeak"])).round(2)

        instance, created = DemandData.objects.get_or_create(GenStationID=39, Date=yesterday)
        instance.Energy = load_factor
        instance.GenType = 'Total'
        instance.save()

    with connection.cursor() as cursor:
        cursor.execute(query_PrevLoadFactor, {'previous_year_day': previous_year_day})
        PrevLoadFactor = cursor.fetchall()

        cursor.close()

    subindex=['i','ii','iii','iv','v','vi','vii','viii']
    alphabets=['a','b','c','d','e','f','g']

    
    # Create the report content as a string
    report_content = f"""
                 TRANSMISSION CORPORATION OF TELANGANA LTD
               GRID OPERATION -- FINAL REPORT FOR {yesterday.strftime('%d/%m/%Y')}
=================================================================================
   |     Generating Station     |Generation at Peak Demand|     Generation  
   |                            |      EX-BUS in MW       |    EX-BUS in MU
Sl |                            |  Morning      Evening   | {yesterday.strftime('%A'):^10}|{previous_year_day.strftime('%A'):^10}
No |                            |  {gridfreq_data.iloc[0,0]:.2f}HZ      {gridfreq_data.iloc[0,1]:.2f}HZ   | {yesterday.strftime('%d/%m/%Y')}|{previous_year_day.strftime('%d/%m/%Y')}
   |   Name            INS.CAP  |  {gridfreq_data.iloc[0,2].strftime('%H:%M')}Hrs     {gridfreq_data.iloc[0,3].strftime('%H:%M')}Hrs  |           |
-----------------------(MW)------------------------------------------------------
(1)     (2)                         (3)         (4)           (5)           (6)
---------------------------------------------------------------------------------
 I  TS GENCO"""

    # Add the data rows to the report content
    for i in range(hydel.shape[0]):
      row_content = f"""
     {hydel.iloc[i,0]:<17}{hydel.iloc[i,2]:>6.0f}{hydel.iloc[i,3]:>12.0f}{hydel.iloc[i,4]:>12.0f}{hydel.iloc[i,5]:>14.3f}    |{hydel.iloc[i,6]:>10.3f}"""
      report_content += row_content

    row_content = f"""
       TS HYDEL-->    {hydel["InstalledCap"].sum():>6.0f}{hydel["MorningPeak"].sum():>12.0f}{hydel["EveningPeak"].sum():>12.0f}{hydel["Energy"].sum():>14.3f}    |{hydel["PrevEnergy"].sum():>10.3f}
        """       
    report_content += row_content


    for i in range(thermal.shape[0]):
      row_content = f"""
     {thermal.iloc[i,0]:<17}{thermal.iloc[i,2]:>6.0f}{thermal.iloc[i,3]:>12.0f}{thermal.iloc[i,4]:>12.0f}{thermal.iloc[i,5]:>14.3f}    |{thermal.iloc[i,6]:>10.3f}"""
      report_content += row_content

    row_content = f"""
       TS THERMAL-->  {thermal["InstalledCap"].sum():>6.0f}{thermal["MorningPeak"].sum():>12.0f}{thermal["EveningPeak"].sum():>12.0f}{thermal["Energy"].sum():>14.3f}    |{thermal["PrevEnergy"].sum():>10.3f}"""       
    report_content += row_content

    row_content = f"""

       TSGENCO TOTAL->{genco["InstalledCap"].sum():>6.0f}{genco["MorningPeak"].sum():>12.0f}{genco["EveningPeak"].sum():>12.0f}{genco["Energy"].sum():>14.3f}    |{genco["PrevEnergy"].sum():>10.3f}"""       
    report_content += row_content

    for i in range(lta.shape[0]):
      row_content = f"""

     {lta.iloc[i,0]:<17}{lta.iloc[i,2]:>6.0f}{lta.iloc[i,3]:>12.0f}{lta.iloc[i,4]:>12.0f}{lta.iloc[i,5]:>14.3f}    |{lta.iloc[i,6]:>10.3f}"""
      report_content += row_content

    row_content = f"""

II CENTRAL SECTOR"""       
    report_content += row_content
    row_content = f"""
     {central_sector.iloc[0,0]:<17}{central_sector.iloc[0,2]:>6.0f}{central_sector.iloc[0,3]:>12.0f}{central_sector.iloc[0,4]:>12.0f}{central_sector.iloc[0,5]:>14.3f}    |{central_sector.iloc[0,6]:>10.3f}"""
    report_content += row_content

    row_content = f"""

III TSShare OF APISGS->{APISGS.iloc[0,2]:>5.0f}{APISGS.iloc[0,3]:>12.0f}{APISGS.iloc[0,4]:>12.0f}{APISGS.iloc[0,5].round(2):>14.3f}    |{APISGS.iloc[0,6]:>10.3f}"""       
    report_content += row_content

    row_content = f"""


                 TRANSMISSION CORPORATION OF TELANGANA LTD
               GRID OPERATION -- FINAL REPORT FOR {yesterday.strftime('%d/%m/%Y')}
==================================================================================
   |     Generating Station     |Generation at Peak Demand|        Generation  
   |                            |      EX-BUS in MW       |       EX-BUS in MU
Sl |                            |  Morning      Evening   |{yesterday.strftime('%A'):^12}|{previous_year_day.strftime('%A'):^10}
No |                            |  {gridfreq_data.iloc[0,0]:.2f}HZ      {gridfreq_data.iloc[0,1]:.2f}HZ   |{yesterday.strftime('%d/%m/%Y'):^12}|{previous_year_day.strftime('%d/%m/%Y'):^10}
   |   Name            INS.CAP  |  {gridfreq_data.iloc[0,2].strftime('%H:%M')}Hrs     {gridfreq_data.iloc[0,3].strftime('%H:%M')}Hrs  |            |
-----------------------(MW)-------------------------------------------------------
(1)    (2)                          (3)          (4)            (5)         (6)
----------------------------------------------------------------------------------
"""
    report_content += row_content

    row_content = f"""
IV  PRIVATE SECTOR

     a) """   

    report_content += row_content
    row_content = f"""{private.iloc[0,0]:<14}{private.iloc[0,2]:>6.0f}{private.iloc[0,3]:>12.0f}{private.iloc[0,4]:>12.0f}{private.iloc[0,5]:>16.3f}   |{private.iloc[0,6]:>8.3f}"""
    report_content += row_content

    row_content = f"""
     b) """       
    report_content += row_content
    for i in range(wind.shape[0]):
      row_content = f"""{wind.iloc[i,0]:<14}{wind.iloc[i,2]:>6.1f}{wind.iloc[i,3]:>12.0f}{wind.iloc[i,4]:>12.0f}{wind.iloc[i,5]:>16.3f}   |{wind.iloc[i,6]:>8.3f}"""
      report_content += row_content

    row_content = f"""

     c) TOTAL SOLAR      {solar["InstalledCap"].sum():>6.2f}"""       
    report_content += row_content
    for i in range(solar.shape[0]):
      row_content = f"""
       {subindex[i]}){' '*(1-i)} {solar.iloc[i,0][:6]:<11}{solar.iloc[i,2]:>6.2f}{solar.iloc[i,3]:>11.0f}{solar.iloc[i,4]:>12.0f}{solar.iloc[i,5]:>16.3f}   |{solar.iloc[i,6]:>8.3f}"""
      report_content += row_content

    row_content = f"""

     d) NONCONVENTIONAL   {nonconventional["InstalledCap"].sum():>5.1f}{' ':12}{'':12}"""       
    report_content += row_content
    for i in range(nonconventional.shape[0]):
        if i==0:
            row_content = f"""
       {subindex[i]}){'  '}{nonconventional.iloc[i,0]:<18}{nonconventional.iloc[i,3]:>11.0f}{nonconventional.iloc[i,4]:>12.0f}{nonconventional.iloc[i,5]:>16.3f}   |{nonconventional.iloc[i,6]:>8.3f}"""
            report_content += row_content
        else:
            row_content = f"""
       {subindex[i]}){' '}{nonconventional.iloc[i,0]:<18}{'':>11}{'':>12}{nonconventional.iloc[i,5]:>16.3f}   |{nonconventional.iloc[i,6]:>8.3f}"""
            report_content += row_content
 
    row_content = f"""

     PRIVATE SECTOR TOTAL   {private_total["MorningPeak"].sum():12.0f}{private_total["EveningPeak"].sum():12.0f}{private_total["Energy"].sum():>16.3f}   |{private_total["PrevEnergy"].sum():>8.3f}"""       
    report_content += row_content

    row_content = f"""

V   STATE PURCHASES   {'':>6}{' ':12}{' ':12}"""       
    report_content += row_content

    for i in range(state_purchases.shape[0]):
      row_content = f"""
     {alphabets[i]}) {state_purchases.iloc[i,0]:<14}{'':>6}{state_purchases.iloc[i,3]:>12.0f}{state_purchases.iloc[i,4]:>12.0f}{state_purchases.iloc[i,5]:>16.3f}   |{state_purchases.iloc[i,6]:>8.3f}"""
      report_content += row_content
    row_content = f"""

     STATE PURCHASE TOTAL   {state_purchases["MorningPeak"].sum():12.0f}{state_purchases["EveningPeak"].sum():12.0f}{state_purchases["Energy"].sum():>16.3f}   |{state_purchases["PrevEnergy"].sum():>8.3f}"""
    report_content += row_content
    row_content = f"""

VI  THIRD PARTY PURCHASES   {third_party_purchase.iloc[0,3]:>12.0f}{third_party_purchase.iloc[0,4]:>12.0f}{third_party_purchase.iloc[0,5]:>16.3f}   |{third_party_purchase.iloc[0,6]:>8.3f}
    (CONSUMER PURCHASES)"""
    report_content += row_content

    row_content = f"""

VII THIRD PARTY SALES       {third_party_sales.iloc[0,3]:>12.0f}{third_party_sales.iloc[0,4]:>12.0f}{third_party_sales.iloc[0,5]:>16.3f}   |{third_party_sales.iloc[0,6]:>8.3f}
    (INTRA STATE PVT GENERATORS)"""
    report_content += row_content

    row_content = f"""

VIII TOTAL DEMAND & CONSUMP {gen_total["MorningPeak"]:>12.0f}{gen_total["EveningPeak"]:>12.0f}{gen_total["Energy"]:>16.3f}   |{gen_total["PrevEnergy"]:>8.3f}
     (WITH PUMPS)"""
    report_content += row_content

    row_content = f"""

IX  {pump.iloc[0,0]:<24}{pump.iloc[0,3]:>12.0f}{pump.iloc[0,4]:>12.0f}{pump.iloc[0,5]:>16.3f}   |{pump.iloc[0,6]:>8.3f}"""
    report_content += row_content

    row_content = f"""


X   {pump.iloc[1,0]:<24}{pump.iloc[1,3]:>12.0f}{pump.iloc[1,4]:>12.0f}{pump.iloc[1,5]:>16.3f}   |{pump.iloc[1,6]:>8.3f}
    """
    report_content += row_content

    row_content = f"""

XI  TS DEMAND(EX-BUS) {gen_total_wo_pump["InstalledCap"]:<6.0f}{gen_total_wo_pump["MorningPeak"]:>12.0f}{gen_total_wo_pump["EveningPeak"]:>12.0f}{'':>16}   |{PrevTSDemand[0][0]:>8.0f}


       ENERGY (MU)    {'':>6}{'':>12}{'':>12}{gen_total_wo_pump["Energy"]:>16.3f}   |{gen_total_wo_pump["PrevEnergy"]:>8.3f}"""
    report_content += row_content

    row_content = f"""


XII  Load Factor      {'':<6}{'':>12}{'':>12}{load_factor:>16.3f}%  |{PrevLoadFactor[0][0]:>8.3f}%
"""
    report_content += row_content


    romannumerals=['XIII','XIV','XV','XVI']
    for i in range(maxcitysolardata.shape[0]):
        row_content = f"""
{romannumerals[i]:<5}{maxcitysolardata.iloc[i,1]:<25}{maxcitysolardata.iloc[i,2]:>5.0f} MW    AT   {str(maxcitysolardata.iloc[i,3])[:5]:>8} Hrs
"""   
        #print(row_content)
        report_content += row_content



    



    row_content = f"""

                             CENTRAL PROJECTS GENERATION (MU)
                       ============================================
                       {'Station':<18}{'Generation':^14}{'Month':^14}
                       {'':<18}{'On Date':^14}{'Cumulative':^14}
                       --------------------------------------------"""       
    report_content += row_content

    for i in range(centralgendata_today.shape[0]):
      row_content = f"""
        {'':<15}{centralgendata_today.iloc[i,1]:<18}{centralgendata_today.iloc[i,2]:>10.3f}{centralgendata_today.iloc[i,4]:>14.3f}"""
      report_content += row_content

    row_content = f"""
                       ============================================"""       
    report_content += row_content

    row_content = f"""

TOTAL SCHEDULES & DRAWALS FROM CENTRAL NETWORK INCLUDING CENTRAL GENERATING STATIONS(MU)
=======================================================================================
 {'State':<12}{'Energy':^12}{'Actual':^12}{'Excess/':^10}{'Cumulative for the Month':^28}{'Cum Excess/':^14}
 {'':^12}{'Scheduled':^12}{'Util.':^12}{'Deficit':^10}{'Share':^14}{'Utilisation':^14}{'Deficit':^14}
---------------------------------------------------------------------------------------"""       
    report_content += row_content

    for i in range(schdrwldata_today.shape[0]):
      row_content = f"""
 {schdrwldata_today.iloc[i,1]:<15}{schdrwldata_today.iloc[i,2]:>8.3f}{schdrwldata_today.iloc[i,3]:>10.3f}{schdrwldata_today.iloc[i,7]:>10.3f}{schdrwldata_today.iloc[i,5]:>14.3f}{schdrwldata_today.iloc[i,6]:>14.3f}{schdrwldata_today.iloc[i,8]:>12.3f}"""
      report_content += row_content

    row_content = f"""
======================================================================================="""       
    report_content += row_content
    lta.set_index('GenStationID',inplace=True)   
    row_content = f"""

    {'GENERATION SUMMARY AS ON '+yesterday.strftime('%d/%m/%Y')+' (MU)':^81}
    {'----------------------------------------':^81}
    {'TS HYDEL GEN .........':<25}{hydel["Energy"].sum():>8.3f}{'':<15}{'TS SHARE of APISGS..':<25}{APISGS.iloc[0,5]:>8.3f}
    {'TS THERMAL GEN........':<25}{thermal["Energy"].sum():>8.3f}{'':<15}{'PRIVATE SECTOR......':<25}{private_total["Energy"].sum():>8.3f}
    {'TS GENCO TOTAL........':<25}{genco["Energy"].sum():>8.3f}{'':<15}{'STATE PURCHASES.....':<25}{state_purchases["Energy"].sum():>8.3f}
    {'SINGARENI... .........':<25}{lta.loc[18,'Energy']:>8.3f}{'':<15}{'3RD PARTY PURC+SALES':<25}{(third_party_purchase.iloc[0,5]+third_party_sales.iloc[0,5]):>8.3f}
    {'CHATTISGARH SPDCL.....':<25}{lta.loc[20,'Energy']:>8.3f}{'':<15}{'SSLB PUMP CONSUMP...':<25}{pump.iloc[0,5]:>8.3f}
    {'CGS UTIL..............':<25}{central_sector.iloc[0,5]:>8.3f}{'':<15}{'NSR PUMP CONSUMP....':<25}{pump.iloc[1,5]:>8.3f}
    {'':<25}{'':>8}{'':<15}{'TOTAL':>25}{gen_total_wo_pump["Energy"]:>8.3f}
    ---------------------------------------------------------------------------------
                                                            THIS YEAR       LAST YEAR
   TS GRID DEMAND for {yesterday.strftime('%d %B')} (in MU){' '*(8-len(yesterday.strftime('%d %B')))}                 :{gen_total_wo_pump["Energy"]:>10.3f}    |{gen_total_wo_pump['PrevEnergy']:>11.3f}
   {'Cumulative for the Month Total (in MU)':<55}:{tsdemand_monthcum[0][0]:>10.3f}    |{gen_data[gen_data['GenStationID']==37][['PrevEnergy']].iloc[0,0]:>11.3f}
   {'Cumulative for the Year Total (in MU) (From 1st April)':<55}:{gen_data[gen_data['GenStationID']==38][['Energy']].iloc[0,0]:>10.3f}    |{gen_data[gen_data['GenStationID']==38][['PrevEnergy']].iloc[0,0]:>11.3f}
    
    
    

    """
    report_content += row_content

#    
    row_content = f"""

  RESERVOIR LEVEL PARTICULARS AS ON {today.strftime('%d/%m/%Y')}
  =================================================================================
  {'RESERVOIR':<13}|{'LAST YEAR':^18}|{'THIS YEAR':^18}|{'LEVEL':^7}|{'EQUI-':^8}| {'FRL':^4} | {'MDDL':^4}
  {'':<13}|{'':^18}|{'':^18}|{'RAISE/':^7}|{'VALENT':^8}| {'':^4} | {'':^4}
  {'':<13}|{'LEVEL':^9}|{'STORAGE':^8}|{'LEVEL':^9}|{'STORAGE':^8}|{'FALL':^7}|{'ENERGY':^8}| {'':>4} | {'':>4}
  {'':<13}|{'(ft)':^9}|{'(Tmc)':^8}|{'(ft)':^9}|{'(Tmc)':^8}|{'OVER':^7}|{'(mu)':^8}| {'(ft)':>4} | {'(ft)':>4}
  {'':<13}|{'':^9}|{'':^8}|{'':^9}|{'':^8}|{'PREV-':^7}|{'':^8}| {'':>4} | {'':>4}
  {'':<13}|{'':^9}|{'':^8}|{'':^9}|{'':^8}|{'IOUS':^7}|{'':^8}| {'':>4} | {'':>4}
  {'':<13}|{'':^9}|{'':^8}|{'':^9}|{'':^8}|{'DAY':^7}|{'':^8}| {'':>4} | {'':>4}
  {'':<13}|{'':^9}|{'':^8}|{'':^9}|{'':^8}|{'(ft)':^7}|{'':^8}| {'':>4} | {'':>4}
  ---------------------------------------------------------------------------------
  {'(1)':^13}|{'(2)':^9}|{'(3)':^8}|{'(4)':^9}|{'(5)':^8}|{'(6)':^7}|{'(7)':^8}| {'(8)':>4} | {'(9)':>4}
  ---------------------------------------------------------------------------------
"""           
    report_content += row_content
    levelstoragedata = levelstoragedata.fillna(0)
    for i in range(levelstoragedata.shape[0]):
      row_content = f"""  {levelstoragedata.iloc[i,1]:<13}|{levelstoragedata.iloc[i,2]:>8.2f} |{levelstoragedata.iloc[i,3]:>7.3f} |{levelstoragedata.iloc[i,4]:>8.2f} |{levelstoragedata.iloc[i,5]:>7.3f} |{levelstoragedata.iloc[i,6]:>6.2f} | {levelstoragedata.iloc[i,7]:>6.3f} | {levelstoragedata.iloc[i,8]:>4.0f} | {levelstoragedata.iloc[i,9]:>4.0f}
  ---------------------------------------------------------------------------------
"""
      report_content += row_content

#    row_content = f"""    ===================================================================================
#"""           
#    report_content += row_content
    row_content = f"""                      
  INFLOWS AND DISCHARGES
    
  {'Inflows in Cusecs @ 06:00 Hrs':^31}        {'Discharges in Cusecs (Avg)':^35}
  ============================           ====================(00 to 24)   (06 to 06)
  1. {'U_JURALA':<19}{inflowsdischargedata.iloc[0,5]:>6}         {'1.':>4}{'U_JURALA':<19}{dis_ujurala.iloc[0,4]:>6}{dis_ujurala.iloc[0,5]:>13}
     {'':<19}{'':>6}         {'':>4}{'L_JURALA':<19}{dis_ljurala.iloc[0,5]:>6}

  2. {'SRISAILAM':<21}{'':>6}       {'2.':>4}{'SRISAILAM':<20}
      {inf_srisailam.iloc[0,1]:<18}{inf_srisailam.iloc[0,5]:>6}         {'':>5}{dis_srisailam.iloc[0,1]:<18}{dis_srisailam.iloc[0,4]:>6}{dis_srisailam.iloc[0,5]:>13}
      {inf_srisailam.iloc[1,1]:<18}{inf_srisailam.iloc[1,5]:>6}         {'':>5}{dis_srisailam.iloc[1,1]:<18}{dis_srisailam.iloc[1,4]:>6}{dis_srisailam.iloc[1,5]:>13}
      {inf_srisailam.iloc[2,1]:<18}{inf_srisailam.iloc[2,5]:>6}         {'':>5}{dis_srisailam.iloc[2,1]:<18}{'':>6}{dis_srisailam.iloc[2,5]:>13}
      {inf_srisailam.iloc[3,1]:<18}{inf_srisailam.iloc[3,5]:>6}         {'':>5}{dis_srisailam.iloc[3,1]:<18}{'':>6}{dis_srisailam.iloc[3,5]:>13}
      {'':<18}{'':>6}         {'':>5}{dis_srisailam.iloc[4,1]:<18}{'':>6}{dis_srisailam.iloc[4,5]:>13}
      {'':<18}{'':>6}         {'':>5}{dis_srisailam.iloc[5,1]:<18}{'':>6}{dis_srisailam.iloc[5,5]:>13}
      {'':<18}{'':>6}         {'':>5}{dis_srisailam.iloc[6,1]:<18}{'':>6}{dis_srisailam.iloc[6,5]:>13}
      {'TOTAL':<18}{inf_srisailam.InfDis06to06.sum():>6}         {'':>5}{'TOTAL':<18}{dis_srisailam.InfDis00to24.sum():>6}{dis_srisailam.InfDis06to06.sum():>13}

  3. {"N' SAGAR":<21}{'':>6}       {'3.':>4}{"N' SAGAR":<12}
      {inf_nsagar.iloc[0,1]:<18}{inf_nsagar.iloc[0,5]:>6}         {'':>5}{dis_nsagar.iloc[0,1]:<18}{'':>6}{dis_nsagar.iloc[0,5]:>13}
      {inf_nsagar.iloc[1,1]:<18}{inf_nsagar.iloc[1,5]:>6}         {'':>5}{dis_nsagar.iloc[1,1]:<18}{'':>6}{dis_nsagar.iloc[1,5]:>13}
      {'Total':<18}{inf_nsagar.InfDis06to06.sum():>6}         {'':>5}{dis_nsagar.iloc[2,1]:<18}{'':>6}{dis_nsagar.iloc[2,5]:>13}
      {'':<18}{'':<6}         {'':>5}{dis_nsagar.iloc[3,1]:<18}{'':>6}{dis_nsagar.iloc[3,5]:>13}
  4. {'Pulichintala':<19}{inf_pulichintala.iloc[0,5]:>6}         {'':>5}{dis_nsagar.iloc[4,1]:<18}{dis_nsagar.iloc[4,4]:>6}{dis_nsagar.iloc[4,5]:>13}
      {'':<18}{'':>6}         {'':>5}{dis_nsagar.iloc[5,1]:<18}{'':>6}{dis_nsagar.iloc[5,5]:>13}
  {otherdata.iloc[0,1]:<18}{otherdata.iloc[0,3]:>6} mts         {'':>5}{dis_nsagar.iloc[6,1]:<18}{'':>6}{dis_nsagar.iloc[6,5]:>13}
  {otherdata.iloc[1,1]:<18}{otherdata.iloc[1,3]:>6} mts         {'':>5}{'Total':<18}{dis_nsagar.InfDis00to24.sum():>6}{dis_nsagar.InfDis06to06.sum():>13}
  {otherdata.iloc[2,1]:<18}{otherdata.iloc[2,3]:>6} mts  
  {otherdata.iloc[3,1]:<18}{otherdata.iloc[3,3]:>6} mts         {'4.':>4}{"Pulichintala":<19}{dis_pulichintala.iloc[0,4]:>6}{dis_pulichintala.iloc[0,5]:>13}


  WEATHER:
      {eval(eval(weatherdata.iloc[0,3]))}
      {eval(eval(weatherdata.iloc[1,3]))}
      {eval(eval(weatherdata.iloc[2,3]))}
        




"""   

    report_content += row_content
    response.write(report_content)



    return response

def export_dailymu_to_text(request):

    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename="DailyMU.txt"'
    yesterday = datetime.datetime.strptime(request.GET['date'], '%Y-%m-%d').date()
    today = yesterday + relativedelta(days=1)
    yesterday_str = f'{yesterday.year}-{yesterday.month:02}-{yesterday.day:02}'
    daybeforeyesterday = today - relativedelta(days=2)
    cur_month = yesterday.month
    cur_year = yesterday.year
    cur_yearmonth = f'{cur_year}-{cur_month:02}'

    monthstartday = f'{cur_year}-{cur_month:02}-01'
    fin_year_startday = f'{cur_year-1}-04-01'

    previous_year_day = yesterday - relativedelta(years=1)
    



    query_monthdataTSDemand = """
               SELECT MorningPeak, EveningPeak, Energy, Date
               FROM dailyreport_DemandData
               WHERE Date BETWEEN %(monthstartday)s AND %(yesterday)s AND GenStationID=53
               """

    query_monthdataTSDemandmu = """
               SELECT MorningPeak, EveningPeak, Energy, Date
               FROM dailyreport_DemandData
               WHERE Date BETWEEN %(monthstartday)s AND %(yesterday)s AND GenStationID=36
               """



    query_month_gendata = f"""WITH month_data AS
               (SELECT GenStationID, GenStationName, GenType, InstalledCap, Energy, STRFTIME('%%Y-%%m', Date) AS month, Date
               FROM dailyreport_DemandData)
               
               SELECT GenStationID, GenStationName, GenType, InstalledCap, Energy, Date
               FROM month_data
               WHERE month='{cur_year}-{cur_month:02}'
               """

    query_month_maxcitysolar = f"""
               WITH month_data AS
               (SELECT PID, Name, MaxDemand, Time, STRFTIME('%%Y-%%m', Date) AS month, Date
               FROM dailyreport_MaxCitySolar)
               
               SELECT PID, Name, MaxDemand, Time, Date
               FROM month_data
               WHERE month='{cur_year}-{cur_month:02}'
               """
                
    with connection.cursor() as cursor:


        cursor.execute(query_monthdataTSDemand, {'yesterday': yesterday, 'monthstartday': monthstartday})
        tsdemand_monthdata = cursor.fetchall()

        cursor.execute(query_monthdataTSDemandmu, {'yesterday': yesterday, 'monthstartday': monthstartday})
        tsdemandmu_monthdata = cursor.fetchall()

        cursor.execute(query_month_gendata,{'cur_year':cur_year,'cur_month':cur_month})
        monthgendata = cursor.fetchall()

        cursor.execute(query_month_maxcitysolar,{'cur_year':cur_year,'cur_month':cur_month})
        monthmaxcitysolardata = cursor.fetchall()

        cursor.close()



        tsdemand_monthdata=pd.DataFrame(tsdemand_monthdata,columns=['MorningPeak', 'EveningPeak', 'Energy', 'Date'])
        tsdemand_monthdata['MaxTSDemand']=tsdemand_monthdata[['MorningPeak', 'EveningPeak']].max(axis=1)

        tsdemandmu_monthdata=pd.DataFrame(tsdemandmu_monthdata,columns=['MorningPeak', 'EveningPeak', 'Energy', 'Date'])

        monthgendata = pd.DataFrame(monthgendata,columns=['GenStationID', 'GenStationName', 'GenType', 'InstalledCap', 'Energy', 'Date'])

        monthmaxcitysolardata = pd.DataFrame(monthmaxcitysolardata,columns=['PID','Name','MaxDemand','Time','Date'])
        maxcitysolardata = monthmaxcitysolardata[monthmaxcitysolardata['Date']==yesterday]


        def monthlyenergyreport(df_allgen,type):
            df_filtered=df_allgen[df_allgen['GenType'].isin(type)]
            #print(df_filtered)

            for GenStationID in df_filtered['GenStationID'].unique():
                df_station=df_filtered.loc[df_filtered['GenStationID']==GenStationID]
                df_station['GenStationName']=list(df_station.loc[df_station['Date']==yesterday_str]['GenStationName'])[0]
                df_filtered.loc[df_filtered['GenStationID']==GenStationID]=df_station
            
            df=df_filtered.loc[df_filtered['GenStationID']==1]
#            print(list(df['Energy']))
            df_report=pd.DataFrame(index=df_filtered['GenStationName'].unique(),columns=['InstalledCap']+[f'{cur_year}-{cur_month:02}-{x:02}' for x in range(1,yesterday.day+1)])
            for date in df_report.columns:
                for gen in df_report.index:
                    try:
                        if date=='InstalledCap':
                            df_report.loc[gen,date]=df_filtered[(df_filtered['GenStationName']==gen)&(df_filtered['Date']==yesterday_str)].iloc[0]['InstalledCap']
                            
                        else:
                            df_report.loc[gen,date]=df_filtered[(df_filtered['Date']==date) & (df_filtered['GenStationName']==gen)].iloc[0]['Energy']

                    except IndexError:
                        pass
            df_report['CUM']=df_report.drop('InstalledCap',axis=1).sum(axis=1,skipna=True)
            df_report['AVG']=df_report.drop(['InstalledCap','CUM'],axis=1).mean(axis=1,skipna=True)
            return df_report

        report_hydel=monthlyenergyreport(monthgendata,['Hydel'])
        #print(report_hydel)
        report_thermal=monthlyenergyreport(monthgendata,['Thermal'])
        report_genco=pd.concat([report_hydel,report_thermal],axis=0).reset_index()
        report_thermal['CapUtil']=report_thermal['CUM']*100000/report_thermal['InstalledCap']/24/yesterday.day
        report_thermal=report_thermal[list(report_thermal.columns[:-2])+[report_thermal.columns[-1],report_thermal.columns[-2]]]
#        print(list(report_thermal.columns[:-2])+[report_thermal.columns[-1],report_thermal.columns[-2]])
        report_lta=monthlyenergyreport(monthgendata,['LTA'])
        report_cgs=monthlyenergyreport(monthgendata,['Central Sector'])
        report_apisgs=monthlyenergyreport(monthgendata,['APISGS'])
        report_solar=monthlyenergyreport(monthgendata,['Private_solar'])
        report_nonconventional=monthlyenergyreport(monthgendata,['Private_Nonconventional'])
        report_statepurchases=monthlyenergyreport(monthgendata,['State Purchases','Third Party Purchases','Third Party Sales','Pump'])

        monthdata_private=monthgendata[monthgendata["GenType"].str.contains("Private") & ~(monthgendata["GenType"].isin(['Private_solar','Private_Nonconventional']))]
        for GenStationID in monthdata_private['GenStationID'].unique():
            df_station=monthdata_private.loc[monthdata_private['GenStationID']==GenStationID]
            df_station['GenStationName']=list(df_station.loc[df_station['Date']==yesterday_str]['GenStationName'])[0]
            monthdata_private.loc[monthdata_private['GenStationID']==GenStationID]=df_station
        report_private = pd.DataFrame(index=monthdata_private['GenStationName'].unique(),columns=['InstalledCap']+[f'{cur_year}-{cur_month:02}-{x:02}' for x in range(1,yesterday.day+1)])
        for date in report_private.columns:
            for gen in report_private.index:
                try:
                    if date=='InstalledCap':
                        report_private.loc[gen,date]=monthdata_private[(monthdata_private['GenStationName']==gen)&(monthdata_private['Date']==yesterday_str)].iloc[0]['InstalledCap']
                    else:
                        report_private.loc[gen,date]=monthdata_private[(monthdata_private['Date']==date) & (monthdata_private['GenStationName']==gen)].iloc[0]['Energy']
                except IndexError:
                    pass

        report_private['CUM']=report_private.drop(['InstalledCap'],axis=1).sum(axis=1,skipna=True)
        report_private['AVG']=report_private.drop(['InstalledCap','CUM'],axis=1).mean(axis=1,skipna=True)
        report_totalprivate=pd.concat([report_solar,report_nonconventional,report_private],axis=0)



    report_content = ''
    row_content = f"""TS TRANSCO GENERATION DAYWISE DATA IN MILLION UNITS FROM   {monthstartday[-2:]}/{monthstartday[-5:-3]}/{monthstartday[:4]}  TO  {yesterday.strftime('%d/%m/%Y')}
"""
    report_content += row_content
    
    heading = ['Station','(MW)']+[f'{x:02}' for x in range(1,yesterday.day+1)]+['CUM','%CAPUTIL','  AVG',]
    report_content += f"""{'-'*(24+6+8+8+8+6*yesterday.day)}
"""
    for i in range(len(heading)):
        if i ==0:
            row_content = f"""{heading[i]:<23}"""
        else:
            if heading[i] in ['CUM','AVG','%CAPUTIL']:
                row_content = f"""{heading[i]:>9}"""
            else:
                row_content = f"""{heading[i]:>6}"""
        report_content += row_content
    report_content += """
"""
    report_content += f"""{'-'*(24+6+8+8+8+6*yesterday.day)}
"""
    spacer=[21,8]+[6]*yesterday.day+[9,16]
    spacer_thermal=[21,8]+[6]*yesterday.day+[9,8,8]
    decimal_spec=['',0]+[2]*yesterday.day+[3,2,2]

    def addcontent(df):
        report_content=''
        df.reset_index(inplace=True)
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                if type(df.iloc[i,j])==str:
                    row_content = f"""{df.iloc[i,j]:<{spacer[j]}}"""
                else:
                    row_content = f"""{df.iloc[i,j]:>{spacer[j]}.{decimal_spec[j]}f}"""
                report_content += row_content
            report_content += """
""" 
        return report_content

    def addcontent_thermal(df):
        report_content=''
        df.reset_index(inplace=True)
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                if type(df.iloc[i,j])==str:
                    row_content = f"""{df.iloc[i,j]:<{spacer_thermal[j]}}"""
                else:
                    row_content = f"""{df.iloc[i,j]:>{spacer_thermal[j]}.{decimal_spec[j]}f}"""
                report_content += row_content
            report_content += """
""" 
        return report_content

    def addcontent_series(s,heading):
        report_content=''
        s[0]=heading
        for i in range(len(s)):
            if type(s[i])==str:
                row_content = f"""{s[i]:<{spacer[i]}}"""
            else:
                row_content = f"""{s[i]:>{spacer[i]}.{decimal_spec[i]}f}"""
            report_content+=row_content
            
        return report_content

    def addcontent_series_thermal(s,heading):
        report_content=''
        s[0]=heading
        for i in range(len(s)):
            if type(s[i])==str:
                row_content = f"""{s[i]:<{spacer_thermal[i]}}"""
            else:
                row_content = f"""{s[i]:>{spacer_thermal[i]}.{decimal_spec[i]}f}"""
            report_content+=row_content
            
        return report_content

    def addcontent_series_private(s,heading):
        report_content=''
        for i in range(len(s)):
            if i==0:
                row_content = f"""{heading:<20}"""
            else:
                if  i==1:
                    if heading=='Private Total':
                        row_content = f"""{'':>9}"""
                    else:
                        row_content = f"""{s[i]:>9.2f}"""
                elif i > (len(s)-3):
                    row_content = f"""{s[i]:>9.3f}"""
                else:
                    row_content = f"""{s[i]:>6.2f}"""
            report_content+=row_content
            
        return report_content

    def addcontent_series2(s,heading):
        report_content=''
        for i in range(len(s)):
            if i==0:
                row_content = f"""{heading:<20}"""
            else:
                if  i==1:
                    row_content = f"""{s[i]:>9.0f}"""
                elif i > (len(s)-4):
                    row_content = f"""{s[i]:>9.3f}"""
                else:
                    row_content = f"""{s[i]:>6.2f}"""
            report_content+=row_content
            
        return report_content



    report_content += f"""{addcontent(report_hydel)}"""
    series=report_hydel.sum(axis=0)
#    print(series)
#    series['%CapUtil']=series['CUM']*100000/series['InstalledCap']/24/yesterday.day
    report_content += f"""{addcontent_series(series,'Total Hydel')}

"""
    report_content += f"""{addcontent_thermal(report_thermal)}"""
    s=report_thermal.sum(axis=0)

    try:
        s[-2]=s[-3]*100000/s[1]/24/yesterday.day
    except:
        s[-1]=np.nan
    report_content += f"""{addcontent_series_thermal(s,'Total Thermal')}

"""
#    print(report_genco.sum(axis=0))
    report_content += f"""{addcontent_series(report_genco.sum(axis=0),'TS GENCO TOTAL')}
"""



    #print(report_lta)
    report_content += f"""{addcontent(report_lta)}"""
    report_content += f"""{addcontent(report_cgs)}"""
    report_content += f"""{addcontent(report_apisgs)}
"""
    report_content += """PRIVATE SECTOR:
"""

    decimal_spec=['',2]+[2]*yesterday.day+[3,2,2]
    report_content += f"""{addcontent(report_private)}"""
    report_solar.loc['TOTAL SOLAR',:]=report_solar.sum(axis=0)
    report_content += f"""{addcontent(report_solar)}"""
    report_nonconventional.reset_index(inplace=True)
    report_content += f"""{addcontent_series(report_nonconventional.sum(axis=0),'Nonconventional')}
"""


    report_totalprivate.reset_index(inplace=True)
    #report_totalprivate['InstalledCap']=np.Nan
    report_content += f"""{addcontent_series(report_totalprivate.sum(axis=0),'Private Total')}
"""
    spacer=[24,5]+[6]*yesterday.day+[9,16]
    report_statepurchases['InstalledCap']=''
    report_content += f"""
STATE PURCHASES:
{addcontent(report_statepurchases)}"""


    report_content += f"""{'':30}{'      '*(yesterday.day):>6}{'':8}{'Max':>9}
"""

    report_content += f"""{'TS DEMAND (MU)':<29}"""
    for i in range(tsdemandmu_monthdata['Energy'].shape[0]):
        row_content = f"""{tsdemandmu_monthdata['Energy'][i]:>6.1f}"""
        report_content += row_content
    report_content+=f"""{tsdemandmu_monthdata['Energy'].sum():>9.3f}"""
    report_content+=f"""{tsdemandmu_monthdata['Energy'].max():>9.3f}"""
    report_content+=f"""{tsdemandmu_monthdata['Energy'].mean():>7.2f}"""

    report_content += f"""
{'TS DEMAND (MW)':<29}"""
        
    for i in range(tsdemand_monthdata['MaxTSDemand'].shape[0]):
        row_content = f"""{tsdemand_monthdata['MaxTSDemand'][i]:>6.0f}"""
        report_content += row_content
    try:
        report_content+=f"""{int(tsdemand_monthdata['MaxTSDemand'].max()):>18}"""
    except:
        report_content+=f"""                {np.nan:>8}"""


    monthmaxcitysolardata['Time']= monthmaxcitysolardata['Time'].astype(str)
    report_maxcitysolardemand = pd.DataFrame(index=monthmaxcitysolardata['Name'].unique(),columns=[f'{cur_year}-{cur_month:02}-{x:02}' for x in range(1,yesterday.day+1)])

    for date in report_maxcitysolardemand.columns:
        for gen in report_maxcitysolardemand.index:
            try:
                report_maxcitysolardemand.loc[gen,date]=monthmaxcitysolardata[(monthmaxcitysolardata['Date'].astype(str)==date) & (monthmaxcitysolardata['Name']==gen)].iloc[0,2]
            except:
                pass


    report_maxcitysolartime = pd.DataFrame(index=monthmaxcitysolardata['Name'].unique(),columns=[f'{cur_year}-{cur_month:02}-{x:02}' for x in range(1,yesterday.day+1)])
    for date in report_maxcitysolartime.columns:
        for gen in report_maxcitysolartime.index:
            try:
                report_maxcitysolartime.loc[gen,date] = monthmaxcitysolardata[(monthmaxcitysolardata['Date'].astype(str)==date) & (monthmaxcitysolardata['Name']==gen)].iloc[0,3][:5]

            except:
                pass

    report_maxcitysolardemand['CUM']=' '
    report_maxcitysolardemand['MAX']=report_maxcitysolardemand.drop(['CUM'],axis=1).max(axis=1,skipna=True)
    report_maxcitysolardemand['AVG']=' '


    report_maxcitysolartime['CUM']=' '
    report_maxcitysolartime['MAX']=' '
    report_maxcitysolartime['AVG']=' '
    report_maxcitysolardemand.reset_index(inplace=True)
    report_maxcitysolardemand.set_index('index',inplace=True)
    report_maxcitysolartime.loc['CITY MAX DEMAND  MET','MAX']=report_maxcitysolartime.loc['CITY MAX DEMAND  MET',report_maxcitysolardemand.reset_index().drop(['CUM','index','MAX','AVG'],axis=1).astype(float).idxmax(axis=1,skipna=True)[0]]
    report_maxcitysolartime.loc['SOLAR MAX DEMAND MET','MAX']=report_maxcitysolartime.loc['SOLAR MAX DEMAND MET',report_maxcitysolardemand.reset_index().drop(['CUM','index','MAX','AVG'],axis=1).astype(float).idxmax(axis=1,skipna=True)[1]]
    report_maxcitysolartime.reset_index(inplace=True)
    report_maxcitysolardemand.reset_index(inplace=True)

    spacer=[27,8]+[6]*yesterday.day+[12,16]
    decimal_spec=['',0]+[2]*yesterday.day+[3,2,2]

    def addcontent_series3(s):

        report_content=''
        for i in range(len(s)):
            if type(s[i])==str:
                row_content = f"""{s[i]:<{spacer[i]}}"""
            else:
                row_content = f"""{s[i]:>{spacer[i]}.0f}"""
            report_content+=row_content
        return report_content

    def addcontent_series4(s):

        report_content=''
        s[0]='@ Time'
        for i in range(len(s)):
            if i==0:
                row_content = f"""{s[i]:<{spacer[i]}}"""
            else:
                row_content = f"""{s[i]:>{spacer[i]}}"""
            report_content+=row_content
        return report_content

    report_content += f"""
{addcontent_series3(report_maxcitysolardemand.iloc[0])}
"""
    report_content += f"""{addcontent_series4(report_maxcitysolartime.iloc[0],)}
"""
    report_content += f"""{addcontent_series3(report_maxcitysolardemand.iloc[1])}
"""
    report_content += f"""{addcontent_series4(report_maxcitysolartime.iloc[1],)}
"""
    report_content += f"""{'-'*(24+6+8+8+8+6*yesterday.day)}
"""

    # Save the report content to the response
    response.write(report_content)

    return response
